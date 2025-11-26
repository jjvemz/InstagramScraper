import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_excel(metadata, comments, platform, filename):
    """
    Exportar datos scrapeados a archivo Excel (.xlsx)
    
    Args:
        metadata: Dict con metadatos del post
        comments: Lista de dicts con comentarios
        platform: Nombre de la plataforma (instagram, tiktok, etc)
        filename: Nombre base del archivo (sin extension)
    
    Returns:
        str: Ruta completa del archivo guardado
    """
    os.makedirs(f"scrape/{platform}", exist_ok=True)
    filepath = os.path.join(f"scrape/{platform}", filename + ".xlsx")

    wb = Workbook()
    ws = wb.active

    # Headers de metadatos
    metadata_headers = list(metadata.keys())
    for col, header in enumerate(metadata_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Valores de metadatos
    for col, header in enumerate(metadata_headers, 1):
        ws.cell(row=2, column=col, value=metadata.get(header, ""))

    # Dejar columnas 15 y 16 vacias
    start_col = 17

    # Headers de comentarios
    if comments:
        comment_headers = list(comments[0].keys())
        for col, header in enumerate(comment_headers, start_col):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos de comentarios
        for row_idx, comment in enumerate(comments, start=3):
            for col, header in enumerate(comment_headers, start_col):
                ws.cell(row=row_idx, column=col, value=comment.get(header, ""))

    # Guardar con manejo de errores
    try:
        wb.save(filepath)
        print(f"[OK] XLSX exportado: {filepath}")
        return filepath
    except Exception as e:
        print(f"Error al guardar Excel: {e}")
        # Intentar con nombre de archivo seguro (sin caracteres especiales)
        safe_filename = re.sub(r'[^\w\-_\. ]', '_', filename)
        safe_filepath = os.path.join(f"scrape/{platform}", safe_filename + ".xlsx")
        try:
            wb.save(safe_filepath)
            print(f"[OK] XLSX exportado con nombre seguro: {safe_filepath}")
            return safe_filepath
        except Exception as e2:
            print(f"Error critico al guardar Excel: {e2}")
            raise